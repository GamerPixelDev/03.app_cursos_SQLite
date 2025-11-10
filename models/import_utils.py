import os
from openpyxl import load_workbook
from models import alumnos, cursos, matriculas

def importar_alumnos_desde_excel(ruta):
    wb = load_workbook(ruta)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]  # saltar cabecera
    total, duplicados = 0, 0
    for fila in filas:
        if not fila or not fila[0]:
            continue
        ok = alumnos.crear_alumno(*fila)
        if ok:
            total += 1
        else:
            duplicados += 1
    return {"entidad": "alumnos", "nuevos": total, "duplicados": duplicados}

def importar_cursos_desde_excel(ruta):
    wb = load_workbook(ruta)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]
    total, duplicados = 0, 0
    for fila in filas:
        if not fila or not fila[0]:
            continue
        ok = cursos.crear_curso(*fila)
        if ok:
            total += 1
        else:
            duplicados += 1
    return {"entidad": "cursos", "nuevos": total, "duplicados": duplicados}

def importar_matriculas_desde_excel(ruta):
    wb = load_workbook(ruta)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]
    total, duplicados = 0, 0
    for fila in filas:
        if not fila or not fila[0]:
            continue
        ok = matriculas.crear_matricula(*fila[:2])
        if ok:
            total += 1
        else:
            duplicados += 1
    return {"entidad": "matrículas", "nuevos": total, "duplicados": duplicados}

