import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from models import alumnos, cursos, matriculas

# Carpeta de salida
EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

#=== ALUMNOS ===
def exportar_alumnos_excel():
    datos = alumnos.obtener_alumnos()
    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    # Encabezados
    columnas = [
        "NIF", "Nombre", "Apellidos", "Localidad", "Código Postal",
        "Teléfono", "Email", "Sexo", "Edad", "Estudios", "Estado Laboral"
    ]
    ws.append(columnas)
    # Estilos de encabezado
    for col in ws[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20
    for i, cell in enumerate(ws[1], start=1):
        ws.cell(row=1, column=i).fill = PatternFill("solid", fgColor="3E64FF")
    # Datos
    for fila in datos:
        ws.append(fila)
    # Ajustar anchos de columna
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    # Guardar archivo
    ruta = os.path.join(EXPORT_DIR, "alumnos.xlsx")
    wb.save(ruta)
    return ruta

#=== CURSOS ===
def exportar_cursos_excel():
    datos = cursos.obtener_cursos()
    if not datos:
        raise ValueError("No hay cursos para exportar.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cursos"
    columnas = [
        "Código", "Nombre", "Fecha Inicio", "Fecha Fin",
        "Lugar", "Modalidad", "Horas", "Responsable"
    ]
    ws.append(columnas)
    # --- formato de cabecera ---
    for col in ws[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.alignment = Alignment(horizontal="center")
        col.fill = PatternFill("solid", fgColor="3E64FF")
    # --- datos ---
    for fila in datos:
        ws.append(fila)
    # --- ajustar anchos ---
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    ruta = os.path.join(EXPORT_DIR, "cursos.xlsx")
    wb.save(ruta)
    return ruta

#=== MATRÍCULAS ===
def exportar_matriculas_excel():
    datos = matriculas.obtener_matriculas()
    if not datos:
        raise ValueError("No hay matrículas para exportar.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Matrículas"
    columnas = [
        "NIF", "Nombre Alumno", "Código Curso", "Curso", "Fecha Insc."
    ]
    ws.append(columnas)
    for col in ws[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.alignment = Alignment(horizontal="center")
        col.fill = PatternFill("solid", fgColor="3E64FF")
    for fila in datos:
        ws.append(fila)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    ruta = os.path.join(EXPORT_DIR, "matriculas.xlsx")
    wb.save(ruta)
    return ruta